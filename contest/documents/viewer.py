import ntpath  # на сервере пути будут не в формате NT
import random
import re
from csv import reader
from io import BytesIO, StringIO
from os.path import join
from re import sub
from tempfile import TemporaryFile

try:
    from aspose import slides as aspose_slides
except ImportError:
    aspose_slides = None
try:
    from aspose import words as aspose_words
except ImportError:
    aspose_words = None

from django.contrib.contenttypes.models import ContentType
from mammoth import convert_to_html
from openpyxl import Workbook, load_workbook
from pygments import highlight
from pygments.formatters import HtmlFormatter
from pygments.lexers import CppLexer
from pylatexenc.latexwalker import LatexMacroNode, LatexWalker
from xls2xlsx import XLS2XLSX
from xlsx2html import xlsx2html

from contest.documents import replaces
from contests.models import Attachment, Contest, Course, Problem


def remove_ppt_watermarks(content):
    content = sub(replaces.PPT_WM_1, replaces.TSPAN, content)
    content = sub(replaces.PPT_WM_2, replaces.BLANK, content)
    content = sub(replaces.PPT_WM_3, replaces.BLANK, content)
    content = content.replace(replaces.PPT_ST_1_BEFORE, replaces.PPT_ST_1_AFTER)
    content = content.replace(replaces.PPT_ST_2_BEFORE, replaces.PPT_ST_2_AFTER)
    content = content.replace(replaces.PPT_ST_3_BEFORE, replaces.PPT_ST_3_AFTER)
    return content


def remove_doc_watermarks(content):
    return sub(replaces.DOC_WM_1, replaces.BLANK, content)


def to_html(attachment):
    result = None, False
    attachment_ext = attachment.extension()
    if attachment_ext in ('.h', '.hpp', '.c', '.cpp'):
        content = attachment.file.read()
        formatter = HtmlFormatter(linenos='inline', wrapcode=True)
        result = highlight(content.decode(errors='replace').replace('\t', ' ' * 4), CppLexer(), formatter), False
    elif attachment_ext in ('.ppt', '.pptx'):
        if aspose_slides is not None:
            ppt_file = aspose_slides.Presentation(attachment.file.path)
            options = aspose_slides.export.HtmlOptions()
            byte_stream = BytesIO()
            ppt_file.save(byte_stream, aspose_slides.export.SaveFormat.HTML, options)
            html_content = str(byte_stream.getvalue(), 'utf-8')
            result = remove_ppt_watermarks(html_content), True
    elif attachment_ext in ('.xls', '.xlsx'):
        if attachment_ext == '.xls':
            temp = TemporaryFile()
            converter = XLS2XLSX(attachment.file.path)
            converter.to_xlsx(temp)
        else:
            temp = attachment.file.path
        html_sheets = dict()
        workbook = load_workbook(temp)
        for i in range(len(workbook.sheetnames)):
            xlsx_file = temp
            current_sheet = StringIO()
            xlsx2html(xlsx_file, current_sheet, locale='en', sheet=i)
            current_sheet.seek(0)
            current_sheet_html = current_sheet.read()
            html_sheets[workbook.sheetnames[i]] = f'<div class="slide table_excel" id="slideslideIface{i+1}"> <hr> <h3 id="fsheet{id(workbook.sheetnames[i])}">{workbook.sheetnames[i]}</h3><hr> ' + current_sheet_html + '</div>'
        sheets = ''.join(f'{value}' for sheet_name, value in html_sheets.items())

        nav = '<ul>{}</ul>'.format(''.join(f'<li class="excel_nav"><a href="#fsheet{id(sheet_name)}">{sheet_name}</a></li>'
                                               for sheet_name in html_sheets.keys()))
        workbook.close()
        result = (sheets + nav), True
    elif attachment_ext in ['.htm', '.html']:
        file = attachment.file.open(mode='rb').read().decode('utf8')
        return file.replace('</style>', '\ntd { overflow: hidden; }</style>'), False
    elif attachment_ext == '.csv':
        temp = TemporaryFile()
        workbook = Workbook()
        worksheet = workbook.active
        with open(attachment.file.path, 'r') as file:
            for row in reader(file):
                worksheet.append(row)
        workbook.save(temp)
        sheet = StringIO()
        xlsx2html(temp, sheet, locale='en')
        sheet.seek(0)
        html_content = sheet.read()
        result = html_content, False
    elif attachment_ext == '.doc':
        if aspose_words is not None:
            doc_file = aspose_words.Document(attachment.file.path)
            byte_stream = BytesIO()
            doc_file.save(byte_stream, aspose_words.SaveFormat.DOCX)
            html_content = convert_to_html(byte_stream).value
            result = remove_doc_watermarks(html_content), False
    elif attachment_ext == '.docx':
        result = convert_to_html(attachment.file.path).value, False
    return result


def unique(str_list):
    return list(dict.fromkeys(str_list))  # можно ведь проще: list(set(str_list))


def get_document_class(content):
    result = ''
    parser = LatexWalker(content)

    for node in parser.get_latex_nodes()[0]:
        if node.nodeType() == LatexMacroNode and node.macroname == 'documentclass':
            result = node.latex_verbatim()
            break

    return result


def get_packages(content):
    result = []
    parser = LatexWalker(content)

    for node in parser.get_latex_nodes()[0]:
        if node.nodeType() == LatexMacroNode and node.macroname == 'usepackage':
            chunk = node.latex_verbatim()
            result.append(chunk)

    return result


def merge_languages(content):
    result = content
    languages = set()
    languages_package = '\\usepackage['

    for chunk in result:
        if chunk.endswith('{babel}'):
            languages_list = chunk.lstrip('\\usepackage').lstrip('[').rstrip('{babel}\n').rstrip(']').split(',')
            for language in languages_list:
                languages.add(language.strip())
            result.remove(chunk)

    for language in languages:
        languages_package += (language + ', ')

    languages_package = languages_package[:len(languages_package) - 2]
    languages_package += ']{babel}'
    result.append(languages_package)

    return result


def get_commands(content, special_names=None):
    result = []
    parser = LatexWalker(content)

    for node in parser.get_latex_nodes()[0]:
        if node.nodeType() == LatexMacroNode and node.macroname not in ['documentclass', 'usepackage']:
            if special_names and node.macroname in special_names.keys():
                regex = r'\{\\\w+} \{[ \\\n\t\w{@}]+}' if special_names[node.macroname] else r'((\w*\{*[\\\w\./]+}*)|(=\d+\w+))'
                pattern = re.compile(r'\\' + node.macroname + regex)
                match = re.search(pattern, content)
                result.append(match.group() if match else None)
            else:
                result.append(node.latex_verbatim())

    return result


def rename_user_commands(content, postfix, required_names):
    command_pattern = re.compile(r'\\newcommand\{\\\w+}')
    user_commands = re.findall(command_pattern, content)
    result = content

    for user_command in user_commands:
        old_command_name = re.search(r'\{\\\w+}', user_command)[0]
        old_command_name = old_command_name.lstrip('{\\').rstrip('}')

        if old_command_name in required_names:
            new_command_name = old_command_name + postfix
            for end in [r'\\', r'\n', r'}', r'\.']:
                name_pattern = r'\\' + old_command_name + end
                result = re.sub(name_pattern, r'\\' + new_command_name + (end if end != r'\.' else '.'), result)

    return result


def get_concat_command(content, required_name):
    result = '\n'

    for chunk in content:
        if chunk.startswith('\\newcommand{\\' + required_name + '}'):
            result += chunk.replace('newcommand', 'renewcommand')
            break

    result = re.sub(r'\n', '\n\t', result) + '\n'

    return result


def get_body(content):
    pattern = re.compile(r'\\begin\{document}.*\\end\{document}', re.DOTALL)
    result = re.search(pattern, content).group().lstrip('\\begin{document}').rstrip('\\end{document}')
    result = re.sub(r'%.*\n', '\n', result)
    result = re.sub(r'\n *\n', '\n', result)
    return result


def problem_replace(content, problems_list, problem_index):
    start = content.find('#!') + 2
    end = content.find('!#', start)
    old = content[start:end]
    new = problems_list[problem_index].description
    content = content.replace('#!' + old + '!#', new, 1)
    return content


def tex_gen(attachment):
    file = str(attachment.file)

    with open(attachment.file.path, encoding='utf-8') as in_file:
        file_content = in_file.read()
        pattern = r'#!(.*?)!#'
        matches = re.findall(pattern, file_content)
        levels = dict((y, x) for x, y in Problem.DIFFICULTY_CHOICES)
        if len(matches) > 0:
            file = join(attachment.dirname, 'tex_gen_temp.tex')  # этот файл ведь после отображения пользователю уже не нужен? если да, тогда лучше его сделать временным

            with open(file, 'w', encoding='utf-8') as out_file:
                for match in matches:
                    error = "Возникла проблема при распознавании фильтра: %s Проверьте его и попробуйте снова."
                    match_components = match.split('/')

                    try:
                        found_course = Course.objects.get(title_official=match_components[0])
                    except Course.DoesNotExist:
                        return file, error.format("Неверно задан курс.")

                    try:
                        found_contest = Contest.objects.get(title=match_components[1], course=found_course)
                    except Contest.DoesNotExist:
                        return file, error.format("Неверно задан раздел.")

                    if match_components[-1].endswith('.tex'):
                        found_file = None

                        if len(match_components) == 3:
                            for attachment in Attachment.objects.filter(object_type=ContentType.objects.get(model='contest'), object_id=found_contest.id):  # у раздела и задачи есть .attachment_set.all()
                                if ntpath.basename(attachment.file.name) == match_components[-1]:
                                    found_file = attachment.file.path
                                    break

                        if len(match_components) == 4:
                            found_problem = Problem.objects.get(title=match_components[-2], contest=found_contest)
                            for attachment in Attachment.objects.filter(object_type=ContentType.objects.get(model='problem'), object_id=found_problem.id):
                                if ntpath.basename(attachment.file.name) == match_components[-1]:
                                    found_file = attachment.file.path
                                    break
                                    
                        if found_file is None:
                            return file, error.format("Неверно задано имя файла.")

                        with open(found_file, 'r', encoding='utf-8') as found_file:
                            found_file_content = found_file.read()
                            file_content = tex_concat(file_content, 'dest', found_file_content, 'from', match)

                    elif match_components[-1] in levels:
                        found_problem = Problem.objects.filter(difficulty=levels[f'{match_components[2]}'], contest=found_contest)  # тут тоже можно использовать обратную реляцию found_contest.problem_set.filter(...)
                        try:
                            random_problem = random.randint(0, len(found_problem) - 1)
                        except ValueError:
                            file_content = re.sub(pattern, '', file_content,
                                                  count=1)
                            continue
                        try:
                            file_content = re.sub(pattern, found_problem[random_problem].description, file_content, count=1)  # а что будет, если в поле description есть html-тэги?
                        except re.error:
                            file_content = problem_replace(file_content, found_problem, random_problem)

                    else:
                        found_problem = Problem.objects.filter(title=match_components[2], contest=found_contest)

                        try:
                            file_content = re.sub(pattern, found_problem[0].description, file_content, count=1)
                        except re.error:
                            file_content = problem_replace(file_content, found_problem, 0)

                file_content = file_content.replace('{{', '{ {').replace('}}', '} }')
                out_file.write(file_content)
                start = file.find('upload') + 7
                file = file[start:].replace('\\', '/')

    return file, None


def tex_concat(dest_content, dest_postfix, from_content, from_postfix, match):
    result = ''
    pattern = r'#!' + match + r'!#'
    rename_command_names = ['theme', 'class', 'header', 'tasks']
    special_command_names = {'columnsep': False, 'newlength': False, 'graphicspath': False, 'everymath': False,
                             # 'makeatletter': False, 'makeatother': False,
                             'newcounter': False, 'DeclareRobustCommand': True}

    dest_content = rename_user_commands(dest_content, dest_postfix, rename_command_names)
    from_content = rename_user_commands(from_content, from_postfix, rename_command_names)

    commands = get_commands(dest_content, special_command_names) + get_commands(from_content, special_command_names)
    packages = get_packages(dest_content) + get_packages(from_content)
    dest_body = get_body(dest_content)
    from_body = get_body(from_content) + get_concat_command(commands, 'zZ')

    commands = unique(commands)
    packages = unique(packages)
    packages = merge_languages(packages)

    result += (get_document_class(dest_content) + '\n\n')

    for package in packages:
        result += (package + '\n')

    for command in commands:
        result += (command + '\n')

    result += '\n\\begin{document}'
    result += dest_body.replace(pattern, from_body)
    result += '\\end{document}\n'

    return result
