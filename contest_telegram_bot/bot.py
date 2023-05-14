import atexit
import json
import mimetypes
import os
import sys
import threading
from copy import deepcopy
from importlib import import_module

import requests
import telebot
from PyPDF2.errors import FileNotDecryptedError
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.http import HttpRequest, QueryDict
from django.utils import timezone
from django.utils.datastructures import MultiValueDict
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
from telebot import custom_filters, types
from telebot.apihelper import ApiTelegramException
from telebot.types import Message

from accounts.models import Account
from contest_telegram_bot.constants import login_btn_text, logout_btn_text
from contest_telegram_bot.keyboards import (back_to_problem_keyboard,
                                            moderator_notification_initial_keyboard, notification_control_keyboard,
                                            problem_detail_keyboard, settings_keyboard, staff_and_moders_start_keyboard,
                                            staff_course_contests_keyboard, staff_course_menu_keyboard,
                                            staff_course_problems_keyboard, staff_course_student_menu_keyboard,
                                            staff_course_students_keyboard, staff_notification_initial_keyboard,
                                            staff_problem_menu_keyboard, start_keyboard_unauthorized,
                                            student_table_keyboard, submission_creation_keyboard,
                                            submission_files_control_texts, submissions_list_keyboard_for_staff,
                                            submissions_list_keyboard_for_students, timer_keyboard)
from contest_telegram_bot.models import TelegramUser, TelegramUserSettings
from contest_telegram_bot.utils import (all_content_types_with_exclude,
                                        cancel_notification_text, check_submission_limit_excess, create_file_from_bytes,
                                        file_chunk_size, get_account_by_tg_id,
                                        get_active_course_users, get_all_faculties_without_mfk__ids,
                                        get_all_study_levels__ids, get_contest_user_by_tg_id, get_course_label,
                                        get_telegram_user, get_user_assignments, is_excel_file, is_schedule_file,
                                        json_get, notify_settings_students_faculties_to_bool, notify_specific_tg_users,
                                        notify_specific_tg_users_by_contest_users, progress_bar, send_notification_text,
                                        tg_authorisation_wrapper)
from contests.api import SubmissionCreateAPI
from contests.forms import AttachmentForm, SubmissionFilesAttachmentMixin
from contests.models import Submission
from schedule.models import Schedule, ScheduleAttachment, current_week_date_from, current_week_date_to

tbot = telebot.TeleBot(settings.BOT_TOKEN)
telegram_users_msg_files_info = {}
notification_info = {}


def remove_webhook():
    try:
        tbot.remove_webhook()
    except Exception:
        pass


if settings.BOT_LISTEN:
    try:
        tbot.set_webhook(f'{settings.CONTEST_DOMAIN}/contest-telegram-bot/{settings.BOT_TOKEN}')
    except Exception:
        pass
    atexit.register(remove_webhook)


def welcome_handler(outer_message: types.Message, welcome_text: str = ", добро пожаловать в систему МГУ Контест!"):
    def send_welcome_message(text: str, message: types.Message, keyboard: types.ReplyKeyboardMarkup):
        text += welcome_text
        tbot.send_message(chat_id=message.chat.id, text=text, reply_markup=keyboard)

    def callback_for_authorized(message: types.Message):
        contest_user = get_telegram_user(message.chat.id).contest_user
        contest_user_account = get_account_by_tg_id(chat_id=message.chat.id)
        start_message_text = f'{get_account_by_tg_id(chat_id=message.chat.id)}'
        if contest_user_account.type == 2 or contest_user_account.type == 3:
            keyboard, _ = staff_and_moders_start_keyboard(staff_contest_user=contest_user,
                                                          for_moders=contest_user_account.type == 2)
        else:
            keyboard, _ = student_table_keyboard(table_type='courses', contest_user=contest_user)

        enter_msg = tbot.send_message(chat_id=message.chat.id, text='Выполняется вход...',
                                      reply_markup=timer_keyboard())
        tbot.delete_message(chat_id=message.chat.id, message_id=enter_msg.id)
        send_welcome_message(text=start_message_text, message=message, keyboard=keyboard)

    def callback_for_unauthorized(message: types.Message):
        start_message_text = message.chat.first_name
        if message.chat.last_name is not None:
            start_message_text += " " + message.chat.last_name
        keyboard = start_keyboard_unauthorized()
        send_welcome_message(text=start_message_text, message=message, keyboard=keyboard)

    all_callbacks_kwargs = {'message': outer_message}
    tg_authorisation_wrapper(chat_id=outer_message.chat.id, authorized_fun=callback_for_authorized,
                             unauthorized_fun=callback_for_unauthorized, auth_fun_args=all_callbacks_kwargs,
                             unauth_fun_args=all_callbacks_kwargs)


@tbot.message_handler(commands=['start'], chat_types=['private'])
def start_handler(outer_message: types.Message):
    welcome_handler(outer_message=outer_message)


@tbot.message_handler(text=login_btn_text, chat_types=['private'])
def login_handler(outer_message: types.Message):
    def callback_for_authorized(message: types.Message):
        tbot.send_message(chat_id=message.chat.id, text=f'Вы уже вошли в систему как '
                                                        f'{get_account_by_tg_id(chat_id=message.chat.id)}.')

    def callback_for_unauthorized(message: types.Message):
        login_bot_message = tbot.send_message(chat_id=message.chat.id, text="Введите логин и пароль следующим "
                                                                            "образом:\n"
                                                                            "<логин>\n<пароль>", reply_markup=None)
        tbot.register_next_step_handler(message=login_bot_message, callback=login_data_handler)

    all_callbacks_kwargs = {'message': outer_message}
    tg_authorisation_wrapper(chat_id=outer_message.chat.id, authorized_fun=callback_for_authorized,
                             unauthorized_fun=callback_for_unauthorized, auth_fun_args=all_callbacks_kwargs,
                             unauth_fun_args=all_callbacks_kwargs)


def login_data_handler(message: types.Message):
    login_data = message.text.split()
    username = login_data[0]
    if len(login_data) >= 2:
        password = login_data[1]
        authorized_user = authenticate(username=username, password=password)
    else:
        authorized_user = None

    if authorized_user is not None:
        TelegramUser.objects.get_or_create(chat_id=message.chat.id, contest_user=authorized_user)
        TelegramUserSettings.objects.get_or_create(contest_user=authorized_user)
        welcome_handler(outer_message=message, welcome_text=', вы успешно авторизовались.')
    else:
        login_err_msg = tbot.send_message(chat_id=message.chat.id, text='Введённые логин или пароль неверны. '
                                                                        'Повторите попытку.')
        tbot.register_next_step_handler(message=login_err_msg, callback=login_data_handler)


def unauth_callback_inline_keyboard(outer_call: types.CallbackQuery, callback_for_authorized):
    def callback_for_unauthorized(call: types.CallbackQuery):
        tbot.edit_message_text(text='Вы не вошли ни в один аккаунт.', chat_id=call.message.chat.id,
                               message_id=call.message.id,
                               reply_markup=None)

    all_callbacks_kwargs = {'call': outer_call}
    tg_authorisation_wrapper(chat_id=outer_call.message.chat.id, authorized_fun=callback_for_authorized,
                             unauthorized_fun=callback_for_unauthorized, auth_fun_args=all_callbacks_kwargs,
                             unauth_fun_args=all_callbacks_kwargs)


def unauth_callback_for_messages(message: Message, callback_for_authorized):
    def callback_for_unauthorized():
        tbot.send_message(chat_id=message.chat.id, text='Вы не вошли ни в один аккаунт.',
                          reply_markup=start_keyboard_unauthorized())

    tg_authorisation_wrapper(chat_id=message.chat.id, authorized_fun=callback_for_authorized,
                             unauthorized_fun=callback_for_unauthorized)


@tbot.message_handler(text=logout_btn_text, chat_types=['private'])
def logout_handler(message: types.Message):
    if get_telegram_user(chat_id=message.chat.id) is not None:
        keyboard = start_keyboard_unauthorized()
        tbot.send_message(chat_id=message.chat.id, text=f"Вы успешно вышли из аккаунта "
                                                        f"{get_account_by_tg_id(chat_id=message.chat.id)}.",
                          reply_markup=keyboard)
        TelegramUser.objects.get(chat_id=message.chat.id).delete()


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'exit')
def logout_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        tbot.send_message(chat_id=call.message.chat.id,
                          text=f'Вы успешно вышли из аккаунта {get_account_by_tg_id(chat_id=call.message.chat.id)}.',
                          reply_markup=start_keyboard_unauthorized())
        try:
            tbot.delete_message(chat_id=call.message.chat.id, message_id=call.message.id)
        except ApiTelegramException:
            pass
        tbot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
        TelegramUser.objects.get(chat_id=call.message.chat.id).delete()

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'get_settings')
def get_settings_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        keyboard, text = settings_keyboard(contest_user=get_telegram_user(chat_id=call.message.chat.id).contest_user)
        tbot.edit_message_text(
            text=text,
            chat_id=call.message.chat.id,
            message_id=call.message.id,
            reply_markup=keyboard
        )

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'sett')
def setting_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        all_settings = TelegramUserSettings.objects.get(
            contest_user=get_contest_user_by_tg_id(chat_id=call.message.chat.id))
        setting_action = json_get(call.data, 'act')
        setting_name = json_get(call.data, 'name')
        if setting_action == 'text':
            tbot.answer_callback_query(callback_query_id=call.id,
                                       text=all_settings._meta.get_field(setting_name).verbose_name,
                                       show_alert=True)
        elif setting_action == 'chg':
            setattr(all_settings, setting_name, not getattr(all_settings, setting_name))
            all_settings.save()
            keyboard, _ = settings_keyboard(
                contest_user=get_contest_user_by_tg_id(chat_id=call.message.chat.id))
            tbot.edit_message_reply_markup(
                chat_id=call.message.chat.id,
                message_id=call.message.id,
                reply_markup=keyboard
            )

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'score')
def score_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        score_value = json_get(call.data, 'value')
        tbot.answer_callback_query(callback_query_id=call.id, text=str(score_value), show_alert=True)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') in ['go', 'back'])
def goback_students_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        keyboard = None
        user = get_telegram_user(chat_id=call.message.chat.id).contest_user
        action_type = json_get(call.data, 'type')
        destination = json_get(call.data, 'to')
        destination_id = int(json_get(call.data, 'id')) if 'id' in json.loads(call.data) else None
        destination_text = None

        if destination == 'course':
            keyboard, destination_text = student_table_keyboard(table_type='contests', contest_user=user,
                                                                table_id=destination_id)
        if action_type == 'go':
            if destination == 'contest':
                keyboard, destination_text = student_table_keyboard(table_type='problems', contest_user=user,
                                                                    table_id=destination_id)
            elif destination == 'problem':
                keyboard, destination_text = problem_detail_keyboard(contest_user=user, problem_id=destination_id)
        else:
            if destination == 'courses':
                keyboard, destination_text = student_table_keyboard(table_type='courses', contest_user=user,
                                                                    table_id=destination_id)
            elif destination == 'contest':
                keyboard, destination_text = student_table_keyboard(table_type='problems', contest_user=user,
                                                                    table_id=destination_id)
            elif destination == 'problem':
                keyboard, destination_text = problem_detail_keyboard(contest_user=user, problem_id=destination_id)
                msg_for_submission_keyboard_deleting = tbot.send_message(chat_id=call.message.chat.id, text='Выход...',
                                                                         reply_markup=timer_keyboard())
                tbot.delete_message(chat_id=call.message.chat.id, message_id=msg_for_submission_keyboard_deleting.id)
        tbot.clear_step_handler(message=call.message)
        tbot.edit_message_text(text=destination_text, chat_id=call.message.chat.id, message_id=call.message.id,
                               parse_mode='HTML', reply_markup=keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') in ['staff_go', 'staff_back'])
def goback_staff_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        keyboard = None
        contest_user = get_telegram_user(chat_id=call.message.chat.id).contest_user
        contest_user_account = get_account_by_tg_id(chat_id=call.message.chat.id)
        action_type = json_get(call.data, 'type')
        destination = json_get(call.data, 'to')
        destination_id = int(json_get(call.data, 'id')) if 'id' in json.loads(call.data) else None
        destination_text = None

        if destination == 'course':
            keyboard, destination_text = staff_course_menu_keyboard(course_id=destination_id)
        elif destination == 'contests':
            keyboard, destination_text = staff_course_contests_keyboard(course_id=destination_id)
        elif destination == 'contest':
            keyboard, destination_text = staff_course_problems_keyboard(contest_id=destination_id)
        elif destination == 'problem':
            keyboard, destination_text = staff_problem_menu_keyboard(problem_id=destination_id)
        elif destination == 'course_students':
            keyboard, destination_text = staff_course_students_keyboard(course_id=destination_id)
        elif destination == 'stud':
            course_id = int(json_get(call.data, 'crs_id'))
            student_id = int(json_get(call.data, 'stu_id'))
            keyboard, destination_text = staff_course_student_menu_keyboard(course_id=course_id, student_id=student_id)

        if action_type == 'staff_go':
            pass
        else:
            if destination == 'courses':
                remove_from_notification_list(notif_creator_tg_id=call.message.chat.id,
                                              notif_settings_msg_id=call.message.id)
                keyboard, destination_text = staff_and_moders_start_keyboard(staff_contest_user=contest_user,
                                                                             for_moders=contest_user_account.type == 2)

        tbot.clear_step_handler(message=call.message)
        tbot.edit_message_text(text=destination_text, parse_mode='HTML', chat_id=call.message.chat.id,
                               message_id=call.message.id,
                               reply_markup=keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'prob_sett')
def staff_problem_menu_settings_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        problem_id = int(json_get(call.data, 'prob_id'))
        cur_val = int(json_get(call.data, 'cur_val'))
        keyboard, _ = staff_problem_menu_keyboard(problem_id=problem_id,
                                                  show_submissions_number=not cur_val)

        tbot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                       message_id=call.message.id,
                                       reply_markup=keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'stu_sett')
def staff_student_menu_settings_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        course_id = int(json_get(call.data, 'crs_id'))
        student_id = int(json_get(call.data, 'stu_id'))
        cur_val = int(json_get(call.data, 'cur_val'))
        keyboard, _ = staff_course_student_menu_keyboard(course_id=course_id,
                                                         student_id=student_id,
                                                         show_submissions_number=not cur_val)
        tbot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                       message_id=call.message.id,
                                       reply_markup=keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'subm_lst')
def submission_list_for_staff(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        user = User.objects.get(pk=json_get(call.data, 'stud_id'))
        problem_id = json_get(call.data, 'id')
        back_to = json_get(call.data, 'from')
        keyboard, text = submissions_list_keyboard_for_staff(contest_user=user,
                                                             problem_id=problem_id,
                                                             back_to=back_to)
        tbot.edit_message_text(text=text, chat_id=call.message.chat.id, message_id=call.message.id, parse_mode='HTML',
                               reply_markup=keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'problem')
def problem_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        user = get_telegram_user(chat_id=call.message.chat.id).contest_user
        problem_id = json_get(call.data, 'id')
        problem_item = json_get(call.data, 'item')
        if problem_item == 'submissions':
            tbot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.id,
                                           reply_markup=submissions_list_keyboard_for_students(contest_user=user,
                                                                                               problem_id=problem_id))

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


def remove_from_notification_list(notif_creator_tg_id: int, notif_settings_msg_id: int):
    if notification_info.get(notif_creator_tg_id) is not None:
        if notification_info[notif_creator_tg_id].get(notif_settings_msg_id):
            del notification_info[notif_creator_tg_id][notif_settings_msg_id]
        if len(notification_info[notif_creator_tg_id]) == 0:
            del notification_info[notif_creator_tg_id]


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'notify')
def staff_and_moders_notify_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        user = get_telegram_user(chat_id=call.message.chat.id).contest_user
        notif_creator = json_get(call.data, 'creator')
        course_id = json_get(call.data, 'course_id')
        if notification_info.get(call.message.chat.id) is None:
            notification_info[call.message.chat.id] = {}
        notification_info[call.message.chat.id][call.message.id] = {'creator': notif_creator}
        cur_notification_info = notification_info[call.message.chat.id][call.message.id]
        if course_id is not None:
            course_id = int(course_id)
            cur_notification_info['for'] = course_id

        if notif_creator == 'staff':
            keyboard, text = staff_notification_initial_keyboard(course_id=course_id)
        else:
            cur_notification_info['for'] = {'moders': {'open': False,
                                                       'faculties': {}},
                                            'staff': {'open': False,
                                                      'faculties': {}},
                                            'stu': {'open': False,
                                                    'faculties': {}}}
            all_faculties = get_all_faculties_without_mfk__ids()
            all_levels = get_all_study_levels__ids()
            cur_notification_info['for']['moders']['faculties'] = all_faculties
            cur_notification_info['for']['staff']['faculties'] = deepcopy(all_faculties)
            for faculty_id in all_faculties:
                cur_notification_info['for']['stu']['faculties'][faculty_id] = {'open': False,
                                                                                'levels': deepcopy(all_levels)}
            keyboard, text = moderator_notification_initial_keyboard(notification_settings=cur_notification_info['for'])

        tbot.edit_message_text(text=text, chat_id=call.message.chat.id,
                               message_id=call.message.id, reply_markup=keyboard, parse_mode='HTML')
        tbot.register_next_step_handler_by_chat_id(chat_id=call.message.chat.id, callback=get_notify_text,
                                                   notification_initial_msg_id=call.message.id, creator_user=user)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'n_set.opn')
def notify_open_setting_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        path_of_obj_to_open = str(json_get(call.data, 'obj')).split('.')
        group_type = path_of_obj_to_open[0]
        cur_notification_info_for = notification_info[call.message.chat.id][call.message.id]['for']
        settings_of_group_type = cur_notification_info_for[group_type]
        if len(path_of_obj_to_open) == 1:
            settings_of_group_type['open'] = not settings_of_group_type['open']
        elif len(path_of_obj_to_open) == 2:
            settings_faculty = settings_of_group_type['faculties']
            faculty_id = int(json_get(call.data, 'f_id'))
            settings_faculty[faculty_id]['open'] = not settings_faculty[faculty_id]['open']

        updated_keyboard, _ = moderator_notification_initial_keyboard(notification_settings=cur_notification_info_for)
        tbot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.id,
                                       reply_markup=updated_keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'n_set.set')
def notify_select_setting_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        path_of_obj_to_open = str(json_get(call.data, 'obj')).split('.')
        group_type = path_of_obj_to_open[0]
        cur_notification_info_for = notification_info[call.message.chat.id][call.message.id]['for']
        settings_of_group_type = cur_notification_info_for[group_type]
        settings_faculty = settings_of_group_type['faculties']

        if len(path_of_obj_to_open) == 1:
            all_faculties_ids = get_all_faculties_without_mfk__ids()
            if group_type != 'stu':
                if len(settings_faculty) == 0:
                    settings_of_group_type['faculties'] = get_all_faculties_without_mfk__ids()
                else:
                    settings_faculty.clear()
            else:
                if notify_settings_students_faculties_to_bool(settings_faculty) == 0:
                    for faculty_id in all_faculties_ids:
                        settings_faculty[faculty_id]['levels'] = get_all_study_levels__ids()
                else:
                    for faculty_id in all_faculties_ids:
                        settings_faculty[faculty_id]['levels'].clear()
        else:
            faculty_id = json_get(call.data, 'f_id')
            if len(path_of_obj_to_open) == 2:
                if group_type != 'stu':
                    if faculty_id in settings_faculty:
                        settings_faculty.remove(faculty_id)
                    else:
                        settings_faculty.append(faculty_id)
                else:
                    this_faculty_settings = settings_faculty[faculty_id]
                    if len(this_faculty_settings['levels']) == 0:
                        this_faculty_settings['levels'] = get_all_study_levels__ids()
                    else:
                        this_faculty_settings['levels'].clear()
            else:
                faculty_level_id = json_get(call.data, 'l_id')
                this_faculty_settings__levels = settings_faculty[faculty_id]['levels']
                if faculty_level_id in this_faculty_settings__levels:
                    this_faculty_settings__levels.remove(faculty_level_id)
                    this_faculty_settings__levels.remove(faculty_level_id + 1)
                else:
                    this_faculty_settings__levels.append(faculty_level_id)
                    this_faculty_settings__levels.append(faculty_level_id + 1)

        updated_keyboard, _ = moderator_notification_initial_keyboard(notification_settings=cur_notification_info_for)
        tbot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.id,
                                       reply_markup=updated_keyboard)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


def get_notify_text(message: Message, notification_initial_msg_id: int, creator_user: User):
    if message.text is None:
        notify_control_keyboard = None
        text = 'В качестве объявления можно отправить только текст.'
        tbot.register_next_step_handler_by_chat_id(chat_id=message.chat.id, callback=get_notify_text,
                                                   notification_initial_msg_id=notification_initial_msg_id,
                                                   creator_user=creator_user)
    else:
        notify_control_keyboard = notification_control_keyboard()
        text = 'Выберите действие.'
        tbot.register_next_step_handler_by_chat_id(chat_id=message.chat.id, callback=send_notify_text,
                                                   notification_initial_msg_id=notification_initial_msg_id,
                                                   creator_user=creator_user, notify_msg=message,
                                                   notify_control_keyboard=notify_control_keyboard)
    tbot.send_message(chat_id=message.chat.id, text=text, reply_markup=notify_control_keyboard)


def send_notify_text(message: Message, notify_msg: Message, notification_initial_msg_id: int,
                     notify_control_keyboard, creator_user: User):
    if message.text not in [send_notification_text, cancel_notification_text]:
        tbot.send_message(chat_id=message.chat.id, text='Сначала завершите процесс отправки сообщения.',
                          reply_markup=notify_control_keyboard)
        tbot.register_next_step_handler_by_chat_id(chat_id=message.chat.id, callback=send_notify_text,
                                                   notification_initial_msg_id=notification_initial_msg_id,
                                                   creator_user=creator_user, notify_msg=message,
                                                   notify_control_keyboard=notify_control_keyboard)
    else:
        def send_message_with_status(status_msg: str, send_message=True):
            sending_msg = tbot.send_message(chat_id=message.chat.id, text=status_msg,
                                            reply_markup=timer_keyboard())
            if send_message:
                notify_specific_tg_users_by_contest_users(
                    notification_msg=f'Оповещение от <b>{get_account_by_tg_id(chat_id=message.chat.id)}</b>\n\n' +
                                     notify_msg.text,
                    contest_users=recipients)
            tbot.delete_message(chat_id=message.chat.id, message_id=sending_msg.id)

        notification_creator = notification_info[message.chat.id][notification_initial_msg_id]['creator']
        notification_for = notification_info[message.chat.id][notification_initial_msg_id]['for']
        if message.text == send_notification_text:
            if isinstance(notification_for, int):
                course_id = int(notification_for)
                _, recipients = get_active_course_users(course_id=course_id)
            else:
                moderators = Account.objects.filter(type=2,
                                                    faculty_id__in=notification_for['moders']['faculties']).exclude(
                    user=creator_user)
                staff = Account.objects.filter(type=3, faculty_id__in=notification_for['staff']['faculties'])
                students = Account.objects.filter(type=-1)
                students_faculties_info = notification_for['stu']['faculties']
                for students_faculty_id in students_faculties_info.keys():
                    students = students.union(Account.objects.filter(type=1,
                                                                     faculty_id=students_faculty_id,
                                                                     level__in=
                                                                     students_faculties_info[students_faculty_id][
                                                                         'levels']))
                recipients = students.union(moderators, staff).values_list('user')

            send_message_with_status(status_msg='Отправка сообщения...')
            tbot.send_message(chat_id=message.chat.id, text='Сообщение успешно отправлено.')
        else:
            send_message_with_status(status_msg='Отмена операции...', send_message=False)
            remove_from_notification_list(notif_creator_tg_id=message.chat.id,
                                          notif_settings_msg_id=notification_initial_msg_id)

        if notification_creator == 'staff':
            keyboard, text = staff_course_menu_keyboard(course_id=int(notification_for))
        else:
            keyboard, text = staff_and_moders_start_keyboard(staff_contest_user=creator_user, for_moders=True)
        tbot.send_message(chat_id=message.chat.id, text=text, parse_mode='HTML', reply_markup=keyboard)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'submission')
def submission_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        contest_user = get_telegram_user(chat_id=call.message.chat.id).contest_user
        problem_id = json_get(call.data, 'id')
        submission_type = json_get(call.data, 'sub_type')
        files_allowed_extensions = SubmissionFilesAttachmentMixin.FILES_ALLOWED_EXTENSIONS
        if check_submission_limit_excess(user=contest_user, problem_id=problem_id):
            tbot.answer_callback_query(callback_query_id=call.id,
                                       text=f'Вы отправили максимальное количество посылок : {get_user_assignments(user=contest_user, problem_id=problem_id).submission_limit}',
                                       show_alert=True)
            return

        if submission_type == 'Files':
            msg_text = f'Отправьте файлы.\nДопустимы следующие форматы:' \
                       f'<code>{", ".join(files_allowed_extensions)}</code>'
        else:
            msg_text = 'Отправьте голосовые сообщения с Вашим ответом.'

        tbot.edit_message_text(text=msg_text,
                               parse_mode='HTML',
                               chat_id=call.message.chat.id,
                               message_id=call.message.id,
                               reply_markup=back_to_problem_keyboard(problem_id=problem_id))
        telegram_users_msg_files_info[call.message.chat.id] = {'submission_type': submission_type,
                                                               'files_messages': [],
                                                               'problem_id': problem_id,
                                                               'contest_user': contest_user}

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.message_handler(func=lambda message: telegram_users_msg_files_info.get(message.chat.id) is not None and
                                           message.text in submission_files_control_texts())
def submission_files_control(message: Message):
    def callback_for_authorized():
        tg_user_msg_files_info = telegram_users_msg_files_info[message.chat.id]
        submission_type = tg_user_msg_files_info['submission_type']
        files_messages_list = tg_user_msg_files_info['files_messages']
        contest_user = tg_user_msg_files_info['contest_user']
        problem_id = tg_user_msg_files_info['problem_id']
        keyboard, done, cancel = submission_creation_keyboard()

        def back_to_problem(back_loading_text: str, back_text: str):
            back_loading_msg = tbot.send_message(chat_id=message.chat.id, text=back_loading_text,
                                                 reply_markup=timer_keyboard())
            tbot.delete_message(chat_id=message.chat.id, message_id=back_loading_msg.id)
            tbot.send_message(chat_id=message.chat.id, text=back_text)
            problem_keyboard, problem_msg_text = problem_detail_keyboard(contest_user=contest_user,
                                                                         problem_id=problem_id)
            tbot.send_message(chat_id=message.chat.id, text=problem_msg_text, reply_markup=problem_keyboard)

        def create_submission():
            new_submission_id = Submission.objects.all().count() + 1
            files = []
            files_streams = []
            for i, file_message in enumerate(files_messages_list):
                def downloading_progress_text(file_progress_chunks: int, file_progress_percentage):
                    return f'Загрузка файла...\n' \
                           f'{progress_bar(loaded_chunks=file_progress_chunks, total_chunks=10)} {file_progress_percentage}%\n' \
                           f'Загружено файлов:\n' \
                           f'{progress_bar(loaded_chunks=int(10 * (i / len(files_messages_list))), total_chunks=10)} ' \
                           f'{i}/{len(files_messages_list)}'

                cur_submission_attachment_filename = f'submission_{new_submission_id}_{i}'
                if submission_type == 'Verbal':
                    message_file = file_message.voice
                    cur_submission_attachment_filename += '.ogg'
                else:
                    if file_message.document is not None:
                        message_file = file_message.document
                        cur_submission_attachment_filename = f'{message_file.file_name}'
                    else:
                        message_file = file_message.photo[-1]
                        cur_submission_attachment_filename += '.jpg'

                progress_message = tbot.send_message(chat_id=message.chat.id,
                                                     text=downloading_progress_text(file_progress_chunks=0,
                                                                                    file_progress_percentage=0),
                                                     reply_to_message_id=file_message.id)
                with requests.get(tbot.get_file_url(message_file.file_id), stream=True) as r:
                    r.raise_for_status()
                    total = int(r.headers['content-length'])
                    chunk_size = file_chunk_size(total)
                    with open(cur_submission_attachment_filename, 'wb') as f:
                        for chunk in r.iter_content(chunk_size=chunk_size):
                            if chunk:
                                f.write(chunk)
                                percentage = f.tell() / total
                                tbot.edit_message_text(
                                    text=downloading_progress_text(file_progress_chunks=int(10 * percentage),
                                                                   file_progress_percentage=f'{100 * percentage: .2f}'),
                                    chat_id=message.chat.id,
                                    message_id=progress_message.id)
                tbot.delete_message(chat_id=message.chat.id, message_id=progress_message.id)

                cur_file = open(cur_submission_attachment_filename, 'rb')
                content_type, charset = mimetypes.guess_type(cur_submission_attachment_filename)
                files.append(InMemoryUploadedFile(file=cur_file, field_name='FileField',
                                                  name=cur_submission_attachment_filename, content_type=content_type,
                                                  size=sys.getsizeof(cur_file), charset=charset))
                files_streams.append(cur_file)

            request = HttpRequest()
            request.method = "POST"
            request.POST = QueryDict(mutable=True)
            engine = import_module(settings.SESSION_ENGINE)
            request.session = engine.SessionStore()
            request.user = contest_user
            request.FILES = MultiValueDict({'files': files})
            response = SubmissionCreateAPI.as_view()(request=request, problem_id=problem_id)
            json_response = json.loads(response.content)

            for file_stream in files_streams:
                filename = file_stream.name
                file_stream.close()
                os.remove(filename)

            if json_response['status'] != 'OK':
                error_message_text = f'<b>{json_response["status"]}.</b>\n'
                errors_list = json_response.get('errors')
                if errors_list is not None:
                    for cur_err_type in errors_list.values():
                        for i, err in enumerate(cur_err_type):
                            error_message_text += f'\n{i+1}) ' + err['message']
                error_message_text += '\n\nВы можете прислать другие файлы или отменить операцию.'

                tbot.send_message(chat_id=message.chat.id, text=error_message_text,
                                  parse_mode='HTML',
                                  reply_markup=keyboard)
                files_messages_list.clear()
                return False
            else:
                return True

        if message.text == cancel:
            back_to_problem(back_loading_text='Отмена операции...', back_text='Операция отменена.')
            telegram_users_msg_files_info.pop(message.chat.id, None)
        elif message.text == done:
            if len(files_messages_list) == 0:
                if submission_type == 'Verbal':
                    msg_filetype = 'голосового сообщения'
                else:
                    msg_filetype = 'файла'
                tbot.send_message(chat_id=message.chat.id,
                                  text=f'Вы не отправили ни одного {msg_filetype}.',
                                  parse_mode='HTML',
                                  reply_to_message_id=message.id)
                return

            if create_submission():
                back_to_problem(back_loading_text='Завершение процесса отправки...',
                                back_text='Посылка успешно отправлена.')
                telegram_users_msg_files_info.pop(message.chat.id, None)

    unauth_callback_for_messages(message=message, callback_for_authorized=callback_for_authorized)


@tbot.message_handler(content_types=all_content_types_with_exclude(),
                      func=lambda message: telegram_users_msg_files_info.get(message.chat.id) is not None)
def submission_file_handler(message: Message):
    def callback_for_authorized():
        def is_correct_file():
            if incorrect_msg_filetype:
                tbot.send_message(chat_id=message.chat.id,
                                  text=f'Это сообщение не является {correct_msg_filetype}.',
                                  parse_mode='HTML',
                                  reply_to_message_id=message.id)
                return False
            return True

        tg_user_msg_files_info = telegram_users_msg_files_info[message.chat.id]
        submission_type = tg_user_msg_files_info['submission_type']
        messages_with_files = tg_user_msg_files_info['files_messages']
        max_files_count = AttachmentForm.FILES_MAX
        keyboard, done, cancel = submission_creation_keyboard()
        incorrect_msg_filetype = False

        if submission_type == 'Verbal':
            if message.voice is None:
                incorrect_msg_filetype = True
                correct_msg_filetype = 'голосовым'
            else:
                msg_filetype_word_plural = 'голосовые сообщения'
        else:
            if message.document is None and message.photo is None:
                incorrect_msg_filetype = True
                correct_msg_filetype = 'документом'
            else:
                msg_filetype_word_plural = 'файлы'

        if not is_correct_file():
            return

        if len(messages_with_files) < max_files_count:
            messages_with_files.append(message)
            if len(messages_with_files) < max_files_count:
                msg_text = f'Вы можете продолжить присылать {msg_filetype_word_plural}.\n' \
                           f'Ещё можно отправить <b>{max_files_count - len(messages_with_files)} файлов.</b>'
        if len(messages_with_files) >= max_files_count:
            msg_text = f'Вы прислали максимальное количество файлов: <b>{max_files_count}.</b>'

        tbot.send_message(chat_id=message.chat.id,
                          text=msg_text,
                          parse_mode='HTML',
                          reply_to_message_id=message.id,
                          reply_markup=keyboard)

    unauth_callback_for_messages(message=message, callback_for_authorized=callback_for_authorized)


@tbot.callback_query_handler(func=lambda call: json_get(call.data, 'type') == 'status')
def status_callback(outer_call: types.CallbackQuery):
    def callback_for_authorized(call: types.CallbackQuery):
        submission_id = int(json_get(call.data, 'status_obj_id'))
        tbot.answer_callback_query(callback_query_id=call.id,
                                   text=f'{Submission.objects.get(pk=submission_id).get_status_display()}',
                                   show_alert=True)

    unauth_callback_inline_keyboard(outer_call=outer_call, callback_for_authorized=callback_for_authorized)


@tbot.channel_post_handler(content_types=['document'],
                           func=lambda message: message.chat.id in settings.SCHEDULE_CHANNELS_IDS)
def schedule_callback(message: Message):
    if is_schedule_file(filename=message.document.file_name) is not None:
        def create_schedule_files(schedule):
            ScheduleAttachment.objects.filter(schedule=schedule).delete()
            schedule.save()
            schedule_file = message.document
            schedule_filename = schedule_file.file_name
            create_file_from_bytes(
                file_bytes=tbot.download_file(file_path=tbot.get_file(schedule_file.file_id).file_path),
                filename=schedule_filename)
            if is_excel_file(message.document.file_name):
                wb = load_workbook(filename=schedule_filename)
                sheets = wb.sheetnames
                wb.close()
                for sheet in sheets:
                    cur_xls = load_workbook(filename=schedule_filename)
                    cur_sheet_filename = f'{sheet}.xlsx'
                    sheets_to_delete = deepcopy(sheets)
                    sheets_to_delete.remove(sheet)
                    for del_sheet in sheets_to_delete:
                        cur_xls.remove(cur_xls[del_sheet])
                    cur_xls.save(cur_sheet_filename)
                    cur_xls.close()
                    with open(cur_sheet_filename, 'rb') as current_course_sch:
                        ScheduleAttachment.objects.update_or_create(schedule=schedule, name=sheet,
                                                                    file=File(current_course_sch))
                    os.remove(cur_sheet_filename)
            else:
                tmp_pdf_file = open(schedule_filename, 'rb')
                pdf_sch_object = PdfReader(tmp_pdf_file)
                if pdf_sch_object.is_encrypted:
                    pdf_sch_object.decrypt('70')  # если пароль сменится - все сломается

                try:
                    for page in list(pdf_sch_object.pages):
                        output = PdfWriter()
                        output.add_page(page)
                        current_course_name = get_course_label(pdf_content=page.extract_text())
                        current_course_filename = f'{current_course_name}.pdf'
                        with open(current_course_filename, "wb") as outputStream:
                            output.write(outputStream)

                        with open(current_course_filename, 'rb') as current_course_sch:
                            ScheduleAttachment.objects.update_or_create(schedule=schedule, name=current_course_name,
                                                                        file=File(current_course_sch))
                        os.remove(current_course_filename)
                except FileNotDecryptedError:
                    tmp_pdf_file.close()
                    os.remove(schedule_filename)
                    return

                tmp_pdf_file.close()
            os.remove(schedule_filename)
            if not same_schedule:
                recipients = TelegramUser.objects.filter(contest_user__is_active=True,
                                                         contest_user__is_superuser=False)
                notify_specific_tg_users(notification_msg=f'{action} <b>расписание на {week} '
                                                          f'({new_schedule.date_from} - {new_schedule.date_to})</b>',
                                         tg_users=recipients, notification_obj=new_schedule)

        current_date = timezone.now()
        if 1 <= current_date.isocalendar()[2] <= 4:
            next_week = False
            week = 'эту неделю'
        else:
            next_week = True
            week = 'следующую неделю'

        try:
            new_schedule = Schedule.objects.get(date_from=current_week_date_from(next_week=next_week),
                                                date_to=current_week_date_to(next_week=next_week))
            schedule_update_time_passed = (timezone.now() - new_schedule.date_updated).seconds
            action = 'Обновлено'
            if schedule_update_time_passed <= 40:
                same_schedule = True
                threading.Timer(40 - schedule_update_time_passed, create_schedule_files, [new_schedule]).start()
            else:
                same_schedule = False
                create_schedule_files(new_schedule)
        except ObjectDoesNotExist:
            new_schedule = Schedule.objects.create(date_created=current_date, date_updated=current_date,
                                                   date_from=current_week_date_from(next_week=next_week),
                                                   date_to=current_week_date_to(next_week=next_week),
                                                   owner=User.objects.get(username='telegram_bot'))
            same_schedule = False
            action = 'Добавлено'
            create_schedule_files(new_schedule)


@tbot.my_chat_member_handler(func=lambda message: message.new_chat_member.status == 'kicked')
def stop_handler(message: types.ChatMemberUpdated):
    if get_telegram_user(chat_id=message.chat.id) is not None:
        TelegramUser.objects.get(chat_id=message.chat.id).delete()


tbot.add_custom_filter(custom_filters.TextMatchFilter())
tbot.add_custom_filter(custom_filters.TextStartsFilter())
